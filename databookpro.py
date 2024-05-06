
from xbot.app import databook as db

import os
import ast
from os.path import dirname,join
import sys
try:
    # 优先使用内置的pandas和numpy
    #REVIEW - 只在Windows测试过，对于Liunx、Macos不一定可行
    arch = 'x86' if '86' in os.environ['PROCESSOR_ARCHITECTURE'] else 'x64'
    sys.path.append(join(dirname(dirname(os.environ['PYTHONPATH'])),f"support_{arch}\cv-engine\site-packages"))
    import pandas as pd
    from pandas.core.dtypes.inference import is_list_like
    import numpy as np
except:
    # 未找到内置的pandas和numpy的情况下，使用用户安装的pandas和numpy
    import pandas as pd
    from pandas.core.dtypes.inference import is_list_like
    import numpy as np

from openpyxl.utils import get_column_letter, get_column_interval, column_index_from_string
import re


'''
TODO
[-] 写入数据时，如果数据是None、''，应该自动转换为None，而且不需要变成字符串格式。
[ ] 支持boolen索引，例如，data = dbp.loc[[True, True, False, True]]
'''

class DataBookBase:
    MAX_NCOLS = 200
    MAX_NROWS = 200_000
    def __init__(self, auto_parse=True, output_format='raw', depth=5):
        self.auto_parse=auto_parse
        self._output_format = output_format
        self.depth = depth

    @property
    def output_format(self):
        return self._output_format
    
    @output_format.setter
    def output_format(self, value):
        if value not in ['raw', 'pandas']:
            raise ValueError('`output_format` can only be set to "raw" or "pandas"'
                             f'but {value} was given')
        self._output_format = value

    @property
    def nrows(self):
        return db.get_row_count()

    @property
    def ncols(self):
        names, descs = self.get_used_nds()
        return len(names)
    
    @property
    def names(self):
        names, descs = self.get_used_nds()
        return names

    def get_used_nds(self):
        '''
        获取被使用的列名和列描述。注意，由于影刀的`databook.get_column_desc_by_name`
        的速度很慢，因此采取了试探性的算法来获取已有列的列信息，所以获取到的列描述有可能
        会不完全。可以通过设置更大的`dbp.depth`来进行更多次试探。
        '''
        #NOTE - 碰壁算法获取列信息
        # 列信息包括列名和列描述。如果获取全部的列描述，然后判断哪些列描述非空，
        # 从而获取有效的信息，要耗时~1s，这是生命不可等待之久。因此我们采取碰壁
        # 算法来获取有效列信息。当连续碰壁`depth`次，我们就会判定后面没有有效的
        # 列描述了。
        # 这种方法存在安全隐患。如果在`depth+1`列有列描述，这些列描述会被遗漏。
        depth_searched = 0
        names = [get_column_letter(i) for i in range(1, db.get_column_count()+1)]
        descs = [db.get_column_desc_by_name(name) for name in names]
        for i in range(len(names)+1, self.MAX_NCOLS+1):
            names.append(get_column_letter(i))
            descs.append(db.get_column_desc_by_name(names[-1]))
            depth_searched = depth_searched+1 if descs[-1] in ['', None] else 0
            if depth_searched > self.depth:    
                return names[:-self.depth-1], descs[:-self.depth-1]
            
    # def get_used_nds(self):
    #     #NOTE - 被遗弃的获取列信息方法
    #     # 该方法首先获取全部的列信息（列名和列描述），然后提取有效的列信息（列描述为空
    #     # 而且该列无数据即为无效列信息）。
    #     # 因为目前影刀提供的获取列描述的函数（databook.get_column_desc_by_name）非常
    #     # 耗时（在我的电脑上，每次获取列描述耗时~0.005秒），如果要获取全部200个列描述，要
    #     # ~1秒，因此该方法被放弃了。
    #     # 如果影刀后续提供了快速的获取全部列描述的接口，那么应该考虑优先使用该方法，
    #     # 以获取完备的列信息
    #     names = [get_column_letter(i) for i in range(1, self.MAX_NCOLS+1)]
    #     descs = [db.get_column_desc_by_name(name) for name in names]
    #     if descs[-1] != '':
    #         end = len(descs)
    #     else:
    #         ends = [i for i in range(-1, -len(descs), -1) if descs[i] == '' and descs[i-1] != '']
    #         end = (len(descs) + ends[0]) if ends else 0
    #     end = max(db.get_column_count(), end)
    #     return names[:end], descs[:end]
            

    @property
    def descs(self):
        names, descs = self.get_used_nds()
        valid_descs = [desc for desc in descs if desc != '']
        if len(valid_descs) != len(set(valid_descs)):
            raise ValueError('Duplicate descriptions are not allowed.')
        return descs
    
    @descs.setter
    def descs(self, value):
        # 检查value的类型和长度
        names, descs = self.get_used_nds()
        if value in [None, '']:
            value = ['']*len(names)
        elif not is_list_like(value):
            raise TypeError(f'Expect list-like of descs, but got `{type(value).__name__}`')
        elif len(value) < len(names):
            value = list(value) + (len(names) - len(value))*['']
        
        # 检查是否有重复列描述
        valid_value = [item for item in value if item not in [None, '']]
        if len(valid_value) != len(set(valid_value)):
            raise ValueError('Duplicate descriptions are not allowed.')

        # 设置列描述
        for i, (item) in enumerate(value):
            name = get_column_letter(i+1)
            db.set_column_desc_by_name(name, item)

    
    def get_all_nds(self):
        names = [get_column_letter(i) for i in range(1, self.MAX_NCOLS+1)]
        descs = [db.get_column_desc_by_name(name) for name in names]
        return names, descs
    
    @property
    def empty(self):
        return self.nrows == 0
    
    @property
    def shape(self):
        return (self.nrows, self.ncols)


class DataBookPro(DataBookBase):
    def __init__(self, auto_parse=True):
        super().__init__(auto_parse)
        self.loc = Location(self, self.auto_parse, self.output_format, self.depth)


    def __getitem__(self, keys):
        return self.loc[:, keys]


    def __setitem__(self, keys, values):
        if isinstance(keys, (str, list)):
            keys = slice(None), keys
        self.loc[keys] = values


    def from_df(self, df):
        '''
        从`pandas.DataFrame`读取数据到数据表格。
        '''
        #TODO - 可选范围
        # 目前会先清空所有数据，然后从第1行、第一列写入数据。
        # 我们应当让用户自己决定要不要清空，从哪行、哪列开始写入数据
        db.clear()
        db.set_range(1, 'A', df.fillna('').values.tolist())
        for i, col in enumerate(df.columns):
            db.set_column_desc_by_name(i+1,col)


    def to_df(self):
        '''
        将数据表格导出为`pandas.DataFrame`。
        '''
        _ = self.output_format
        self.output_format = 'pandas'
        res = self.loc[:, :]
        self.output_format = _
        return res

    
    def clear(self):
        '''清空数据和列备注'''
        db.clear()
        self.descs = None


    def clear_descs(self):
        '''清空全部列备注'''
        self.descs = None


    def __delitem__(self, key):
        names, descs = self.get_used_nds()
        if key not in names and key not in descs:
            raise KeyError(key)
        self.loc[:, key] = ''


    def drop(self, labels=None, axis=0, index=None, columns=None,):
        '''从行或列中删除指定的标签。'''
        if labels is None and index is None and columns is None == 3:
            raise ValueError("Need to specify at least one of 'labels', 'index' or 'columns'")
        elif labels is not None and (index is not None or columns is not None):
            raise ValueError("Cannot specify both 'labels' and 'index'/'columns'")

        if labels is not None:                
            if axis == 0:
                index = labels
            elif axis == 1:
                columns = labels
            else: 
                raise ValueError(f'No axis named {axis} for databook')
        if index is not None:
            self.loc[index, :] = ''
        if columns is not None:
            self.loc[:, columns] = ''


class Location(DataBookBase):
    def __init__(self, dbp, *args, **kwargs):
        self.dbp = dbp
        super().__init__(*args, **kwargs)
        

    def _check_keys_type(self, keys):
        '''检查数据类型。'''
        if callable(keys):
            keys = keys(self.dbp.loc[:, :])

        if isinstance(keys, (int, list, slice, np.ndarray)):
            keys = keys, slice(None)
        elif isinstance(keys, pd.Series):
            keys = keys.values, slice(None)
        elif not isinstance(keys, tuple):
            raise TypeError(f'`{type(keys).__name__}` is not supported for index.')

        if len(keys) > 2:
            raise IndexError('Too many indexers')
        key0, key1 = keys
        if callable(key0):
            key0 = key0(self.dbp.loc[:, :])
        if callable(key1):
            key1 = key1(self.dbp.loc[:, :])

        if not isinstance(key0, (int, slice, list, np.ndarray, pd.Series)):
            raise TypeError(f'`{type(key0).__name__}` is not supported for index')
        if getattr(key0, 'dtype', int) not in [int, bool]:
            raise TypeError(f'`{key0.dtype}` indexes are not supported')
        if not isinstance(key1, (str, slice, list, np.ndarray, pd.Series)):
            raise TypeError(f'`{type(key0).__name__}` is not supported for columns indexer')
        if isinstance(key0, pd.Series):
            key0 = key0.values
        if isinstance(key1, pd.Series):
            key1 = key1.values
        return key0, key1
    

    def _check_key0(self, key0):
        # TODO 
        # - 长度检查，列不能超过200，行不能超过20w？

        if isinstance(key0, slice):
            if key0.step in [1, None]:
                if not isinstance(key0.start, (int, type(None))):
                    raise TypeError(f'Error type(`{type(key0.start).__name__}`) for slice start')
                if not isinstance(key0.stop, (int, type(None))):
                    raise TypeError(f'Error type(`{type(key0.start).__name__}`) for slice stop')
                idxs = slice(key0.start, key0.stop, 1)
            else:
                idxs = self._slc2idx(key0)
        elif isinstance(key0, (int, list, np.ndarray)):
            key0 = [key0] if isinstance(key0, int) else key0
            if max(key0) > self.MAX_NROWS:
                raise KeyError('maximum number of rows exceeded.')
            idxs = np.array(key0)
            if idxs.dtype == int:
                idxs[idxs<0] += self.nrows + 1
                if (idxs <= 0).any():
                    raise KeyError(f'Invalid index: {key0}')

                # FIXME 仅在写入数据时，检查是否有重复的index
                # if len(idxs) != len(set(idxs)):
                #     raise KeyError('Duplicate indices are not allowed.')

            elif idxs.dtype == bool:
                if len(idxs.shape) != 1:
                    raise KeyError('Two dimensional indexes are not supported')
                if idxs.shape[0] > self.MAX_NROWS:
                    raise KeyError('maximum number of rows exceeded.')
                idxs = np.where(idxs)[0]+1


            # databook 不接受numpy的int32，int64类型的数据，所以我们把它转换为整数列表
            idxs = idxs.tolist()
        else:
            raise TypeError(f'`{type(key0).__name__}` is not supported for indices')
        return idxs
    

    def _check_key1(self, key1):
        old_names, old_descs = self.get_used_nds()
        new_descs, new_names = [], []
        names, descs = [], []
        key1 = [key1] if isinstance(key1, str) else key1
        if isinstance(key1, slice):
            if key1.step in [1, None]:
                start, stop = key1.start, key1.stop
                if start == '' or stop == '':
                    raise KeyError('Empty value for slice start/stop')
                names = slice(start, stop, 1)
                descs = None
            else:
                names, descs = self._slc2nds(key1)
        elif isinstance(key1, (list, np.ndarray)):
            key1 = np.array(key1)
            if key1.dtype == bool:
                if len(key1.shape) != 1:
                    raise KeyError('Two dimensional indexes are not supported')
                if key1.shape[0] > self.MAX_NCOLS:
                    raise KeyError('maximum number of columns exceeded.')
                idxs = np.where(key1)[0]+1
                names = [get_column_letter(i) for i in idxs]
                descs = [db.get_column_desc_by_name(name) for name in names]
            else:
                for key in key1:
                    if not isinstance(key, str):
                        raise KeyError('Only str and slice are valid indices')
                    elif key == '':
                        raise KeyError('Empty value for indices')
                    elif re.match('^[A-Z]+$', key):
                        name = key
                        desc = db.get_column_desc_by_name(key)
                        if name not in old_names:
                            old_descs.append(desc)
                            old_names.append(name)
                    else:
                        desc = key
                        if key not in old_descs:
                            # 生成一个新的name
                            # 我们要规避掉已经在使用的name和未来被预订的name。
                            i = len(old_names)+1
                            while True:
                                name = get_column_letter(i)
                                if name in key1:
                                    i+=1
                                else:
                                    break
                            new_descs.append(key)
                            new_names.append(name)
                            old_descs.append(desc)
                            old_names.append(name)
                        else:
                            name = db.get_column_name_by_desc(key)
                            # 检查是否重复赋值
                            if name in key1:
                                raise ValueError(f'Duplicate key({name}, {key}) for names or descriptions are not allowed.')
                    descs.append(desc)
                    names.append(name)
        else:
            raise TypeError(f'`{type(key1).__name__}` is not supported for indices')
        if isinstance(names, list) and names and max(names) > get_column_letter(self.MAX_NCOLS+1):
            raise KeyError('maximum number of cols exceeded.')
        return names, descs, new_names, new_descs


    def _infer_slc_idxs(self, idxs, shape):
        # 检查idxs是否匹配value
        if idxs.start is None and idxs.stop is None:
            row_start = 1
            if shape[0] == 1:
                row_stop = 1 if self.empty else self.nrows
            else:
                row_stop = shape[0]
                if self.nrows != shape[0] and not self.empty:
                    raise ValueError(f'Length of values({shape[0]}) does not match length '
                        f'of index ({self.nrows})')
        elif idxs.start is None:
            row_start = 1
            row_stop = idxs.stop if idxs.stop > 0 else idxs.stop + self.nrows + 1
            nrows = row_stop - row_start + 1
            if shape[0] != 1:
                if row_stop != shape[0]:
                    raise ValueError(f'Length of values({shape[0]}) does not match length '
                        f'of index ({nrows})')
        elif idxs.stop is None:
            row_start = idxs.start if idxs.start > 0 else self.nrows + idxs.start + 1
            row_start = row_start if row_start > 0 else 1
            if shape[0] == 1:
                row_stop = 1 if self.empty else self.nrows            
            else:
                row_stop = row_start + shape[0] - 1
        else:
            if idxs.stop > 0 and idxs.start < 0:
                # 考虑dbp.loc[-2:5]的情况，由于stop已知，而且为正数，所以应该根据stop逆推start。
                row_start = idxs.stop + idxs.start + 1
            else:
                row_start = idxs.start if idxs.start > 0 else self.nrows + idxs.start + 1
            row_start = 1 if row_start <= 0 else row_start
            row_stop = idxs.stop if idxs.stop > 0 else idxs.stop + self.nrows + 1
        nrows = row_stop - row_start + 1
        return row_start, row_stop, nrows


    def _parse_name(self, value):
        if re.match('^[A-Z]+$', value):
            return value
        else:
            name = db.get_column_name_by_desc(value)
            if name:
                return name
            else:
                raise KeyError('["{value}"] not in columns')

    def _infer_slc_names(self, names, shape):
        # 检查nds是否匹配value
        old_names, old_descs = self.get_used_nds()
        start, stop = names.start, names.stop
        if start is None and stop is None:
            name_start = 'A'
            if shape[1] == 1:
                if len(old_names) == 0:
                    name_stop = 'A'
                else:
                    name_stop = old_names[-1]
                ncols = column_index_from_string(name_stop) - column_index_from_string(name_start) + 1
            else:
                if len(old_names) == 0:
                    name_stop = get_column_letter(shape[1])
                else:
                    name_stop = old_names[-1]
                ncols = column_index_from_string(name_stop) - column_index_from_string(name_start) + 1
                if ncols != shape[1]:
                    raise ValueError(f'Dimensions of data({shape[1]}) are inconsistent '
                                     f'with the number of columns({ncols})')
        elif start is None:
            name_start = 'A'
            name_stop = self._parse_name(stop)
            ncols = column_index_from_string(name_stop) - column_index_from_string(name_start) + 1
            if shape[1] != 1 and ncols != shape[1]:
                raise ValueError(f'Dimensions of data({shape[1]}) are inconsistent '
                    f'with the number of columns({ncols})')
        elif stop is None:
            name_start = self._parse_name(start)
            if shape[1] != 1:
                name_stop = get_column_letter(column_index_from_string(name_start)+shape[1]-1)
                ncols = shape[1]
            else:
                name_stop = old_names[-1]
                ncols = column_index_from_string(name_stop) - column_index_from_string(name_start) + 1
        else:
            name_start = self._parse_name(start)
            name_stop = self._parse_name(stop)
            ncols = column_index_from_string(name_stop) - column_index_from_string(name_start) + 1
            if shape[1] != 1 and ncols != shape[1]:
                raise ValueError(f'Dimensions of data({shape[1]}) are inconsistent '
                    f'with the number of columns({ncols})')
        return name_start, name_stop, ncols


    def __setitem__(self, keys, values):
        # 检查keys类型，并转换为列名和列备注
        key0, key1 = self._check_keys_type(keys)
        idxs = self._check_key0(key0)
        names, descs, new_names, new_descs = self._check_key1(key1)

        # 行或者列是空的话，直接返回
        if (isinstance(names, list) and len(names) == 0) or\
            (isinstance(idxs, list) and len(idxs) == 0):
            return

        if isinstance(idxs, slice) and isinstance(names, slice):
            self._setitem_frame(idxs, names, values)
        elif isinstance(idxs, slice):
            self._setitem_columns(idxs, names, values)
        elif isinstance(names, slice):
            self._setitem_rows(idxs, names, values)
        else:
            self._setitem_elemwise(idxs, names, values)
        
        # 设置新的列描述
        for name, desc in zip(new_names, new_descs):
            db.set_column_desc_by_name(name, desc)


    def _setitem_frame(self, idxs, names, values):
        '''以frame的方式写入数据'''
        values = self._preprocess_values(values)

        # 推导行、列起始终止位置
        row_start, row_stop, nrows = self._infer_slc_idxs(idxs, values.shape)
        name_start, name_stop, ncols = self._infer_slc_names(names, values.shape)
        if ncols <= 0 or nrows <= 0:
            return
        
        # 推导values
        if values.shape[0] == 1 and values.shape[1] == 1:
            values = values.repeat(nrows*ncols)
        elif values.shape[0] == 1 and values.shape[1] != 1:
            values = values.repeat(nrows, axis=0)
        values = values.reshape(nrows, ncols).tolist()

        # 写入数据
        db.set_range(row_start, name_start, values)


    def _setitem_columns(self, idxs, names, values):
        '''以列的方式写入数据'''
        # 单列和多列的逻辑有些不一样，考虑如下的代码。我们生成一个3行2列的DataFrame，
        # >>> df = pd.DataFrame(np.arange(6).reshape(3, 2), columns=['a', 'b'])
        # 设置单列：
        # >>> df.loc[:, ['a']] = [1, 2, 3]
        # 设置多列：
        # >>> df.loc[:, ['a', 'b']] = [1, 2]
        # values都是一维列表，但是解析的方式不一样，分别解析为：[[1], [2], [3]]
        # 和[[1, 2], [1, 2], [1, 2]]
        # 设置单列单值
        # >>> df.loc[:, 'A'] = [1]
        # 数据表格有3行，设置单值的时候，会自动推导为：[[1], [1], [1]]
        # 设置多列单值
        # >>> df.loc[:, ['A', 'B']] = [1]
        # 会为每一列都自动推导为：[[1], [1], [1]]
    
        ncols = len(names)
        values = self._preprocess_values(values, ncols=ncols)

        # 对于行索引为切片，如果只有一列，那么需要将values转置
        # dbp.loc[:, 'A'] = [1, 2, 3]
        # values被转换为: [['1', '2', '3']]
        # 这时候values.shape为(1, 3), 我们需要转置其为(3, 1)的数组
        values = values.T if ncols == 1 and values.shape[0] == 1 else values

        # 推导行索引
        row_start, row_stop, nrows = self._infer_slc_idxs(idxs, values.shape)
        if nrows <= 0:
            return

        # 推导values的内容
        if values.shape[0] == 1 and values.shape[1] == 1:
            values = values.repeat(ncols*nrows)
        elif values.shape[0] == 1:
            values = values.repeat(nrows, axis=1)
        elif values.shape[1] > 1:
            values = values.T
        values = values.reshape(ncols, nrows, -1).tolist()

        # 将数据赋值到databook
        for name, item in zip(names, values):
            db.set_range(row_start, name, item)

    
    def _setitem_rows(self, idxs, names, values):
        nrows, ncols = len(idxs), None
        values = self._preprocess_values(values, nrows, ncols)
        name_start, name_stop, ncols = self._infer_slc_names(names, values.shape)
        if ncols <= 0:
            return
        if values.shape[0] == 1 and values.shape[1] == 1:
            values = values.repeat(ncols*nrows)
        elif values.shape[0] == 1 and values.shape[1] != 1:
            values = values.repeat(nrows, axis=0)
        values = values.reshape(nrows, 1, ncols).tolist()
        for idx, item in zip(idxs, values):
            db.set_range(idx, name_start, item)
    

    def _setitem_elemwise(self, idxs, names, values):
        nrows, ncols = len(idxs), len(names)
        # 预处理values。将元素转换为字符串。
        values = self._preprocess_values(values, nrows, ncols)
        
        # 推导values
        if values.shape[0] == 1 and values.shape[1] == 1:
            values = values.repeat(ncols*nrows)
        elif values.shape[0] == 1:
            if nrows != 1 and ncols == 1:
                values = values.T
            elif nrows == 1 and ncols != 1:
                pass
            elif nrows != 1 and ncols != 1:
                values = values.repeat(nrows, axis=0)
        values = values.reshape(nrows, ncols).tolist()

        # 逐元素写入数据
        for idx, row in zip(idxs, values):
            for name, item in zip(names, row):
                db.set_cell(idx, name, item)


    def _preprocess_values(self, values, nrows=None, ncols=None):
        '''
        预处理values。
        将元素转换为字符串，并将values转换为二维列表。
        '''
        if nrows == 1 and ncols == 1:
            values = '' if values is None else str(values)
        elif is_list_like(values):
            shape = np.shape(values)
            if len(shape) > 1 and (nrows == 1 or ncols==1) and (shape[0] != 1):
                values = ['' if item is None else str(item) for item in values]
            elif len(shape) > 2 and (nrows > 1 or ncols > 1):
                values = [['' if item is None else str(item) for item in row] for row in values]
        values = np.array(values)
        values = values.reshape(1, -1) if len(values.shape) < 2 else values
        return values
    

    def _slc2nds(self, slc):
        start = slc.start if slc.start else 'A'
        stop = slc.stop if slc.stop else get_column_letter(self.ncols)
        step = slc.step if slc.step else 1
        if not isinstance(start, str) or\
            not isinstance(stop, str):
            raise TypeError('Error type for column\'s indexer')
        if not isinstance(step, int):
            raise TypeError('Error type for slice step')
        
        name_start = self._parse_name(start)
        name_stop = self._parse_name(stop)
        i_start = column_index_from_string(name_start)
        i_stop = column_index_from_string(name_stop)
        i_stop += 1 if i_stop > i_start else -1
        names = [get_column_letter(i) for i in range(i_start, i_stop, step)]
        descs = [db.get_column_desc_by_name(name) for name in names]
        return names, descs


    def _slc2idx(self, slc):
        step = 1 if slc.step is None else slc.step
        if step > 0:
            start = 1 if slc.start is None else slc.start
            stop = (db.get_row_count() if slc.stop is None else slc.stop) + 1
        elif step < 0:
            start =  db.get_row_count() if slc.start is None else slc.start
            stop = 0 if slc.stop is None else slc.stop -1
        else:
            raise ValueError('slice step cannot be zero')
        start += 0 if start >= 0 else db.get_row_count()+1
        stop += 0 if stop >= 0 else db.get_row_count()+1        
        idxs = np.arange(start, stop, step)
        idxs = idxs[idxs>0].tolist()
        return idxs


    def __getitem__(self, keys):
        key0, key1 = self._check_keys_type(keys)
        idxs = self._check_key0(key0)
        names, descs, new_names, new_descs = self._check_key1(key1)

        # 获取数据
        if isinstance(idxs, slice) and isinstance(names, slice):
            row_start, row_stop, nrows = self._infer_slc_idxs(idxs, (1, 1))
            name_start, name_stop, ncols = self._infer_slc_names(names, (1, 1))
            data = np.array(db.get_range(row_start, name_start, row_stop, name_stop))
        elif isinstance(idxs, slice):
            row_start, row_stop, nrows = self._infer_slc_idxs(idxs, (1, 1))
            data = np.array([db.get_range(row_start, name, row_stop, name) for name in names])
            data = data.T[0]
        elif isinstance(names, slice):
            name_start, name_stop, ncols = self._infer_slc_names(names, (1, 1))
            data = np.array([db.get_range(idx, name_start, idx, name_stop) for idx in idxs])
            data = data[:, 0]
        else:
            data = np.array([[db.get_cell(idx, name) for name in names] for idx in idxs])

        # 自动解析元素。
        if self.dbp.auto_parse:
            raw_data = data
            data = np.empty(data.shape, object)
            for i, row in enumerate(raw_data):
                for j, item in enumerate(row):
                    data[i, j] = self._parse_value(item)

        # 转换数据格式
        names, descs = self._slc2nds(key1) if isinstance(key1, slice) else (names, descs)
        cols = [desc if desc else name for desc, name in zip(descs, names)]
        if isinstance(key0, slice) and isinstance(key1, str):
            data = data[:, 0]
            if self.dbp.output_format == 'pandas':
                data = pd.Series(data, index=self._slc2idx(idxs), name=cols[0])
        elif isinstance(key0, int) and isinstance(key1, slice):
            data = data[0]
            if self.dbp.output_format == 'pandas':
                data = pd.Series(data, index=cols, name=key0)
        elif isinstance(key0, int) and isinstance(key1, str):
            return data[0, 0]
        else:
            if self.dbp.output_format == 'pandas':
                idxs = self._slc2idx(idxs) if isinstance(idxs, slice) else idxs
                data = pd.DataFrame(data, columns=cols, index=idxs)
        return data if self.dbp.output_format == 'pandas' else data.tolist()
    

    def _parse_value(self, expr):
        '''
        解析元素。由于 databook 只能保存字符串，所以我们为用户提供了解析字符串的功能。
        '''
        if expr == '':
            value = None
        else:
            try:
                # 直接解析元素值存在安全隐患，所以我们使用了`ast.literal_eval`来解析 Python 
                # 字面量（包括数字、字符串、元组、列表、字典和布尔值等）。
                value = ast.literal_eval(expr)
            except:
                value = expr
        return value


dbp = DataBookPro()

__all__ = ['dbp']